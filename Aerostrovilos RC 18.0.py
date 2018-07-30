from tkinter import *
from tkinter import ttk
from math import *
import cmath
from matplotlib import pyplot as plt
import matplotlib
from matplotlib.patches import Arc
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from PIL import ImageGrab
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from string import ascii_uppercase
from openpyxl.worksheet.table import Table, TableStyleInfo


win=Tk()
win.wm_title("Aerostrovilos RC 18.0")
icon_2 = PhotoImage(file = "logo.png")
win.call('wm','iconphoto',win._w,icon_2)
win.geometry('900x600')

def calculate():
    m1r=m1rs.get()
    gamma=cp.get()/cv.get()
    ao1=sqrt(gamma*(r.get())*(to1.get()))
    if radiom1rs.get()==0:
        b=calcb1s(m1r,gamma)
        b1s.set(b)
        b=b*pi/180
        calcfunc(m1r,gamma,b)
        f=func.get()
        calck(m.get(),kb1.get(),n.get(),gamma,f,po1.get(),ao1)
        calr1s(k.get(),r1h.get())
        M1.set(m1r*cos(b))
        a1=ao1/sqrt(1+(gamma-1)/2*M1.get()*M1.get())
        c1m.set(M1.get()*a1)
        rv.set(r1h.get()/r1s.get())
        #optimize()
    if radiom1rs.get()==1 and menum1rs.get()=="Set r1h":
        setr1h(ao1,gamma,r1h.get())
    if radiom1rs.get()==1 and menum1rs.get()=="Set rv":
        setrv(ao1,gamma,rv.get())
    t1=to1.get()/(1+(gamma-1)/2*M1.get()*M1.get())
    calculate2(M1.get(),t1)
    if menuw2w1s.get()=='Only Impeller':
        impeller()
    else:
        impeller_with_vanelessspace()
    plotfigure()
    roundoff()

def roundoff():
    roundlist2=["d1s","d1rms","d1h","u1s","u1rms","u1h","w1s","w1rms","w1h","b1s","b1h","b1rms",
    "c1m","to1rel","u2","c2u","c2","c2m","w2","alpha2","dhimp","dhrcl","dhdf","beta2","d2","b2",
    "effi","effstt","effsts","W3","LWRATIO","ARRVD","EFFRVD","W5","ARAD","EFFAD"]
    roundlist4=["m1rs","m1r_rms","m1rh","M1","deffbl","defftc","deffsf","deffrcl","deffdf","defflk",
    "M2","M2U","lid2","rv","d1sd2","M4","M6","cpactrvd","cpactad"]

    for name2 in roundlist2:
        exec("%s.set(round(%s.get(),2))"%(name2,name2))
    for name4 in roundlist4:
        exec("%s.set(round(%s.get(),4))"%(name4,name4))

def export():
    try:
        wb=load_workbook("data.xlsx")
    except:
        wb=Workbook()
    ws=wb.active
    ws.append(["","","","","","","INPUTS"])
    c=ws["G%d"%ws.max_row]
    font1=Font(bold=True,size=12)
    font2=Font(bold=True,size=11)
    c.font=font1#to convert text to bold
    for i in ascii_uppercase[0:12]:
        ws.column_dimensions[i].width=15

    ws.append(["CV\n(J/kg/K)","CP\n(J/kg/K)","R","Pressure Ratio","Mass Flow\n(kg/s)",
    "To1\n(K)","Po1\n(pascal)","N(rpm)","Relative Mach\nnumber shroud","Hub Radius\n(m)",
    "Hub/shroud\nradius ratio","Blockge\nFactor","apha1\n(deg)"])
    ws.append([cv.get(),cp.get(),r.get(),pr.get(),m.get(),to1.get(),po1.get(),n.get(),m1rs.get()
    ,r1h.get(),cv.get(),kb1.get(),alpha1.get()])
    tab=Table(displayName="table1",ref="A2:M3")
    style=TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo=style
    ws.append([""])#empty line

    ws.append(["B2b\n(deg)","W2/W1s","Tip\nClearance","Cf","Blockage\nFactor 2","Vaneless Space\nRadius Ratio",
    "cfvld","r4/r3","Zrd","cfrvd","L/W ratio","Zad","cfad"])
    ws.append([b2b.get(),w2w1s.get(),tipc.get(),cf.get(),kb2.get(),rr32.get(),cfvld.get(),
    r4r3.get(),zrd.get(),cfrvd.get(),lwratio.get(),zad.get(),cfad.get()])
    ws.append([""])#empty line

    ws.append(["","","","RESULTS"])
    c=ws["D%d"%ws.max_row]
    c.font=font1

    ws.append(["","","","Impeller Entry"])
    c=ws["D%d"%ws.max_row]
    c.font=font2

    ws.append(["","Values(mm)","U1(m/s)","W1(m/s)","beta1(deg)","M1 rel","C1m(m/s)"])
    ws.append(["d1s",d1s.get(),u1s.get(),w1s.get(),b1s.get(),m1rs.get(),c1m.get()])
    ws.append(["d1rms",d1rms.get(),u1rms.get(),w1rms.get(),b1rms.get(),m1r_rms.get(),c1m.get()])
    ws.append(["d1h",d1h.get(),u1h.get(),w1h.get(),b1h.get(),m1rh.get(),c1m.get()])
    ws.append(["M1",M1.get(),"Radius Ratio",rv.get()])
    ws.append([""])#empty line

    ws.append(["","","","Impeller Exit"])
    c=ws["D%d"%ws.max_row]
    c.font=font2


    ws.append(["b2\n(mm)","d2\n(mm)","U2\n(m/s)","C2U\n(m/s)","C2\n(m/s)","C2m\n(m/s)","W2(m/s)"])
    ws.append([b2.get(),d2.get(),u2.get(),c2u.get(),c2.get(),c2m.get(),w2.get()])
    ws.append([""])
    ws.append(["M2","M2U","alpha2\n(deg)",u"\u03B1""2\n(deg)","d1s/d2","LI/d2"])
    ws.append([M2.get(),M2U.get(),alpha2.get(),beta2.get(),d1sd2.get(),lid2.get()])
    ws.append([""])

    ws.append(["","","","Impeller Overall"])
    c=ws["D%d"%ws.max_row]
    c.font=font2
    ws.append(["deffbl","defftc","deffsf","deffdf","deffrcl","defflk","Efficiency"])
    ws.append([deffbl.get(),defftc.get(),deffsf.get(),deffdf.get(),deffrcl.get(),defflk.get(),effi.get()])
    ws.append([""])
    ws.append(["DHimp(J)",dhimp.get(),"DHrcl(J)",dhrcl.get(),"DHdf(J)",dhdf.get()])
    ws.append([""])
    if menuw2w1s.get()!='Only Impeller':

        ws.append(["","","","Radial Diffuser"])
        c=ws["D%d"%ws.max_row]
        c.font=font2
        ws.append(["width\n(mm)","len/width","Area Ratio","M4","Efficiency","Cp actual"])
        ws.append([W3.get(),LWRATIO.get(),ARRVD.get(),M4.get(),EFFRVD.get(),cpactrvd.get()])
        ws.append([""])

        ws.append(["","","","Axial Diffuser"])
        c=ws["D%d"%ws.max_row]
        c.font=font2
        ws.append(["width\n(mm)","Area Ratio","M6","Efficiency","Cp actual"])
        ws.append([W5.get(),ARAD.get(),M6.get(),EFFAD.get(),cpactad.get()])
        ws.append([""])

        ws.append(["Overall Compressor"])
        c=ws["A%d"%ws.max_row]
        c.font=font2
        ws.append(["EFFSTT","EFFSTS"])
        ws.append([effstt.get(),effsts.get()])

    wb.save("data.xlsx")
    canvas.print_png("plot.png")

def impeller_with_vanelessspace():
    ## Compressor Mean line Design
    # Input Fluid properties
    R=287
    Nu=1.57*10**-5
    Cv= cp.get()-R
    G=1.4
    G1=(G-1)
    G2=G1/2
    G3=G1/G
    G4=G+1
    G5=G4/(2*G1)
    pi=22/7
    # Input Compressor details from cycle
    EFFSTT=1
    # Inputs for inducer calculations
    Alpha0=0
    N=n.get()
    tb1h=1.8
    tb1s=1
    ZI=16
    Kb1=kb1.get()
    r1h=d1h.get()/(2*1000)
    omega=(2*pi*N)/60
    # Inputs for impeller outlet design
    B2b=b2b.get()*pi/180
    W2W1s=w2w1s.get()
    Tipc=tipc.get()
    Kb2=kb2.get()
    krI=2*10**-4
    # beta1s calculation
    #b1s=(((3+(G*m1rs.get()**2))/(2*m1rs.get()**2))*(1-((1-((4*m1rs.get()**2)/(3+(G*m1rs.get()**2))**2))**0.5)))**0.5
    B1s=b1s.get()*pi/180
    # f(M1rt,B1s) calculation
    CB1s=cmath.cos(B1s)
    SB1s=cmath.sin(B1s)
    # r1h/r1s calculation
    ao1=cmath.sqrt(G*R*to1.get())
    r1hr1s=rv.get()
    r1s=r1h/r1hr1s
    d1s=r1s*2*1000
    # Inlet state conditions calculations
    M1=m1rs.get()*CB1s
    rhoo1=po1.get()/(R*to1.get())
    M1term=1+(G2*(M1**2))
    a1=ao1/(M1term**0.5)
    rho1=rhoo1/(M1term**2.5)
    T1=to1.get()/M1term
    P1=po1.get()*((T1/to1.get())**(G/G1))
    # Inlet velocity triangles calculation
    d1rms=(((d1s**2)+(d1h.get()**2))/2)**0.5
    r1rms=d1rms/(2*1000)
    r1mean=(r1s+r1h)/2
    C1m=M1*a1
    U1s=(2*pi*r1s*N)/60
    U1rms=(2*pi*r1rms*N)/60
    U1h=(2*pi*r1h*N)/60
    W1s=cmath.sqrt((U1s**2)+(C1m**2))
    W1rms=cmath.sqrt((U1rms**2)+(C1m**2))
    W1h=cmath.sqrt((U1h**2)+(C1m**2))
    B1rms=cmath.acos(C1m/W1rms)
    B1h=cmath.acos(C1m/W1h)
    M1rrms=M1/cmath.cos(B1rms)
    M1rh=M1/cmath.cos(B1h)
    b1=r1s-r1h
    # Inlet Throat Area and choke limit estimation
    s1h=(pi*d1h.get())/ZI
    O1h=(s1h*cmath.cos(B1h))-(tb1h*cmath.cos(B1h))
    s1s=(pi*d1s)/ZI
    O1s=(s1s*cmath.cos(B1s))-(tb1s*cmath.cos(B1s))
    Area1th=((b1*1000)*(O1h+O1s)/2)*ZI
    mc=(Area1th*10**-6)*rhoo1*ao1*(((2+G1*((U1rms**2)/(ao1**2)))/G4)**(G4/(2*G1)))
    mcmd=mc/m.get()
    # compressor ideal enthalpy and Specific speed calculation
    # Ns=(N*sqrt(md/rhoo1))/((DHid/9.81)**(3/4))
    # ns=Ns/52.9
    # Stage efficiency loop
    count1=0
    x=0
    while count1 < 1:
        DHid=cp.get()*to1.get()*((pr.get()**G3)-1)
        DHact=DHid/EFFSTT
        To2=to1.get()+(DHact/cp.get())
        # Impeller exit state calculation begins from here
        rho2est=rho1*(To2/to1.get())**(1/G1)
        rho2=rho2est
        DHimp=DHact
        #slipfactor calculation
        CslipU2=cmath.sqrt(cmath.cos(B2b))/(ZI**0.7)
        # Delta H impeller calculations loop
        count2=0
        x1=0
        while count2 < 1:
            #W2/W1s condition check
            C2mU2=0
            W2W1scal=0.00
            while W2W1scal <= W2W1s:
                C2mU2=C2mU2+0.00001
                U2=cmath.sqrt(DHimp/(1-CslipU2-(C2mU2*cmath.tan(B2b))))
                C2m=C2mU2*U2
                C2u=DHimp/U2
                W2u=U2-C2u
                W2=cmath.sqrt((W2u**2)+(C2m**2))
                W2W1scal=W2/W1s
                C2=cmath.sqrt((C2u**2)+(C2m**2))
                Alpha2=cmath.acos(C2m/C2)
                d2_=(U2*60)/(pi*N)
                r2=d2_/2
                Re2=(U2*(d2_/2))/Nu
                # Disk friction loss calculation
                DHdf=(0.01356*rho2*(U2**3)*(d2_**2))/(m.get()*(Re2**0.2))
                # Diffusion factor calculation
                DHimpU22=DHimp/(U2**2)
                D=1-W2W1s+(0.75*DHimpU22*W2W1s*(((ZI/pi)*(1-(r1s/r2))+(2*r1s/r2))**-1))
                # Recirculation loss calculation
                DHrcl=0.02*cmath.sqrt(cmath.tan(Alpha2))*(D**2)*(U2**2)
                # Leakage loss calculation
                b2_=m.get()/(pi*d2_*rho2*C2m*Kb2) # impeller exit width (b2) calculation
                Rc=((2*r2)-(r1h+r1s))/2 # radius of curvature calculation
                Lm=(90/360)*2*pi*Rc  #mean strem line length
                Rbar=(r1mean+r2)/2
                bbar=(b1+b2_)/2
                DPcl=(m.get()*(r2*C2u))/(ZI*Rbar*bbar*Lm)
                Ucl=0.816*cmath.sqrt((2*DPcl)/rho2)
                mcl=rho2*ZI*(Tipc/1000)*Lm*Ucl
                DHlk=(mcl*Ucl*U2)/(2*m.get())
                # New Delta H impeller calculation
                DHimp=DHact-DHdf-DHrcl-DHlk
                complex1=cmath.polar(W2W1scal)
                W2W1scal=complex1[0]

            # Diffusion and blade loading loss (Jansen and Coppage)
            DHbl=0.05*(D**2)*(U2**2)
            # Tip clearence loss calculations (Jansen)
            DHtc=(U2**2)*(0.6*((Tipc/1000)/b2_)*C2mU2*cmath.sqrt(((4*pi)/(b2_*ZI))*(((r1s**2)-(r1h**2))/((r2-r1s)*(1+(rho2/rho1))))*(C2u/U2)*(C1m/U2)))
            # Impeller Skin friction loss calculations starts
            # Impeller axial length calculation
            LI=d2_*cmath.sqrt(0.28*(m1rs.get()+0.8)*(1-(r1rms/r2))*((r1s-r1h)/r2))
            # Impeller Skin friction loss calculations cont..
            Lh=(pi/4)*((LI-(b2_/2))+(r2-r1s-(b1/2)))
            Dh=0.5*(((4*pi*r2*b2_)/((2*pi*r2)+(ZI*b2_)))+((2*pi*((r1s**2)-(r1h**2)))/((pi*(r1s-r1h))+(ZI*b1))))
            Wavg=(W2+((W1s+W1h)/2))/2
            # ####### Calulation of cf in impeller
            Reimp=(Wavg*Dh)/Nu
            Cfsp=0.0625/((cmath.log((krI/3.7*Dh)-(5/Reimp)*cmath.log((krI/3.7*Dh)-(5/Reimp)*cmath.log((krI/3.7*Dh)))))**2)
            Cfcp=Cfsp*(1+(0.075*(Reimp**0.25)*cmath.sqrt(Dh/(2*Rc))))
            Cfrc=Cfcp*((Reimp*((r2/Rc)**2))**0.05)
            DHsf=(Cfrc*(Lh/Dh)*(Wavg**2))+(((W1s**2)*2*r2)/(ZI*Rc))
            #DHsf=((4*Cfrc*Lh*(Wavg**2))/(2*Dh))
            # Impeller efficiency calcultaions
            EFFI=(DHimp-(DHbl+DHtc+DHsf))/DHimp
            # Density calculation
            Po2=po1.get()*((EFFI*((To2/to1.get())-1)+1)**(G/G1))
            T2=To2-((C2**2)/(2*cp.get()))
            P2=Po2*((T2/To2)**(G/G1))
            # New density calculations
            temp2=P2/(R*T2)
            err2=abs((rho2-temp2)/rho2)
            complex2=cmath.polar(err2)
            err2=complex2[0]
            if err2 < 0.00001:
                count2=10
                rho2=temp2
            rho2=temp2
            x1+=1
        if x==0:
            print("rho2=",rho2)
        B2=cmath.acos(C2m/W2)
        a2=cmath.sqrt(G*R*T2)
        M2_=C2/a2
        ao2=cmath.sqrt(G*R*To2)
        MU=U2/ao1
        rhoo2=Po2/(R*To2)
        #----------------------------------------------------------------
        #                   Vaneless space calculations
        #----------------------------------------------------------------
        #Inputs for Vaneless space
        r3=r2*rr32.get()
        r=1
        B0=1

        Ma=M2_
        Alpha=Alpha2
        #Intial Calcultions
        DR=(rr32.get()-1)/10 # dr step size
        ZETA=cfvld.get()*(r2/b2_)
        fx=[None]*11
        fx[0]=(Ma**3)/((1+(G2*(Ma**2)))**G5)
        SS=0
        for i in range(1,11):
            Ma1=Ma
            Alpha1=Alpha
            DS=DR*(r2/cmath.cos(Alpha))
            SS=SS+DS #ds summation
            DELTASTAR=0.037*(SS**-0.2)*((C2/Nu)**-0.2)*DS
            BLOC=B0-2*(DELTASTAR/b2_)
            DB=B0-BLOC
            # delta Msquar calculation step1
            DMSQ=-2*((1+(G2*(Ma**2)))/((Ma**2)-((1/cmath.cos(Alpha))**2)))*((G*(Ma**2)-(cmath.tan(Alpha))**2)*(ZETA/(B0*cmath.cos(Alpha)))+((1/B0)*(DB/DR))-(((1/cmath.cos(Alpha))**2)/r))*(Ma**2)*DR
            # delta TanAlpha calculation step1
            DTALP=(((1/cmath.cos(Alpha))**2)/((Ma**2)-((1/cmath.cos(Alpha))**2)))*((1+G1*(Ma**2))*((ZETA/(B0*cmath.cos(Alpha)))+((1/B0)*(DB/DR))-((Ma**2)/r)))*cmath.tan(Alpha)*DR
            # Assigning the calculated values to new variables
            DMSQ1=DMSQ
            DTALP1=DTALP
            BLOC1=BLOC
            BLOC=B0
            SS=SS-DS
            Ma=cmath.sqrt((Ma1**2)+DMSQ)
            TAlpha=cmath.tan(Alpha1)+DTALP
            Alpha=cmath.atan(TAlpha)
            DS=DR*(r2/cmath.cos(Alpha))
            r=r+DR
            SS=SS+DS
            DELTASTAR=0.037*(SS**-0.2)*((C2/Nu)**-0.2)*DS
            BLOC=B0-2*(DELTASTAR/b2_)
            DB=B0-BLOC
            # delta Msquar calculation step2
            DMSQ=-2*((1+(G2*(Ma**2)))/((Ma**2)-((1/cmath.cos(Alpha))**2)))*((G*(Ma**2)-(cmath.tan(Alpha))**2)*(ZETA/(B0*cmath.cos(Alpha)))+((1/B0)*(DB/DR))-(((1/cmath.cos(Alpha))**2)/r))*(Ma**2)*DR
            # delta TanAlpha calculation step2
            DTALP=(((1/cmath.cos(Alpha))**2)/((Ma**2)-((1/cmath.cos(Alpha))**2)))*((1+G1*(Ma**2))*((ZETA/(B0*cmath.cos(Alpha)))+((1/B0)*(DB/DR))-((Ma**2)/r)))*cmath.tan(Alpha)*DR
            DMSQ=(DMSQ+DMSQ1)/2
            DTALP=(DTALP+DTALP1)/2
            BLOC=(BLOC+BLOC1)/2
            Ma=cmath.sqrt((Ma1**2)+DMSQ)
            TAlpha=cmath.tan(Alpha1)+DTALP
            Alpha=cmath.atan(TAlpha)
            B0=BLOC
            aao=cmath.sqrt(1/(1+(G2*Ma**2)))
            rro=1/((1+(G2*Ma**2))**(1/G1))
            fx[i]=((Ma**3)*aao*rro)

        # Integral calculation using Simpson rule
        INTh=(r3-r2)/10
        INTEGRAL=(fx[0]+fx[10]+(4*(fx[1]+fx[3]+fx[5]+fx[7]+fx[9]))+(2*(fx[2]+fx[4]+fx[6]+fx[8])))*(INTh/3)
        # total pressure loss calculation
        M3=Ma
        Alpha3=Alpha
        Po3=Po2*(1/(1+((G*cfvld.get())/cmath.cos(Alpha2))*(r2/b2_)*(INTEGRAL/(M3*aao*rro))))
        # Static pressure and temperature calculation
        P3=Po3/((1+(G2*M3**2))**(G/G1))
        T3=To2/((Po3/P3)**G3)
        # Vaneless space loss
        DHvld=To2*cp.get()*(((P3/Po3)**G3)-((P3/Po2)**G3))
        if x==0:
            print("DHvld=",DHvld)

        #-----------------------------------------------------------------
        #          Radial vaned diffuser based on Radius Ratio
        #-----------------------------------------------------------------
        # RVD Inputs
        b4=b2_
        # Calculations
        r4=r3*r4r3.get()
        Alp5b=pi/2-Alpha3
        Alp4b=pi/2-cmath.asin(cmath.sin(pi/2+Alp5b)*(r3/r4))
        Alpha4b=pi/2-Alp4b
        # Throat calculation
        w3=(2*pi*r3*cmath.sin(Alp5b))/zrd.get()
        w4=(2*pi*r4*cmath.sin(Alp4b))/zrd.get()
        # RVD Lenght calculation
        Lrvd=(r3*cmath.sin(Alp4b-Alp5b))/(cmath.sin(pi/2-Alp4b))
        # RVD Lenght to throat ratio
        LWratio=Lrvd/w3
        # RVD Area Ratio
        ARrvd=w4/w3
        # RVD CPideal
        CPidrvd=1-(w3/w4)**2
        # RVD semi cone angle
        Thetarvd=cmath.asin(((w4-w3)/2)/Lrvd)
        # RVD Efficiency
        EFFrvd=1/(1+(1.2*cfrvd.get()*(((ARrvd**2)-1)/(2*cmath.tan(Thetarvd)))))
        # RVD CP actual
        CPactrvd=EFFrvd*CPidrvd
        # Static pressure at outlet of RVD
        P4=P3+(CPactrvd*(Po3-P3))
        #Static enthalpy rise in RVD
        Delhrvd=((cp.get()*T3)/EFFrvd)*(((P4/P3)**(G3))-1)
        T4=(Delhrvd/cp.get())+T3
        Po4=P4*((To2/T4)**(G/G1))
        a4=cmath.sqrt(G*R*T4)
        rho4=P4/(R*T4)
        # Mach no at outlet of RVD
        m4=cmath.sqrt((2/G1)*((To2/T4)-1))
        C4=m4*a4
        C4m=m.get()/(rho4*2*pi*r4*b4*B0)
        # Flow angle at exit of RVD
        Alpha4=cmath.acos(C4m/C4)
        # RVD loss estimation
        DHrvd=To2*cp.get()*(((P4/Po4)**G3)-((P4/Po3)**G3))

        #----------------------------------------------------------------
        #                  Axial vaned diffuser based on LWratio
        #---------------------------------------------------------------

        Po5=Po4
        P5=P4
        T5=T4
        r5=r4+0.005
        rho5=rho4
        C5m=C4m
        LWratio=4
        ANUA=m.get()/(rho5*C5m)
        r6=cmath.sqrt((ANUA/pi)+(r5**2))
        ANUH=r6-r5

        Alpha5=Alpha4
        Alp5b=pi/2-Alpha5
        Alp6b=pi/2
        Alpha6b=pi/2-Alp6b
        w5=(2*pi*((r6+r5)/2)*cmath.sin(Alp5b))/zad.get()
        w6=(2*pi*((r6+r5)/2)*cmath.sin(Alp6b))/zad.get()
        #Lrd=(r5*sin(Alp4b-Alp3b))/(sin(pi/2-Alp4b))
        Lad=lwratio.get()*w5
        ARad=w6/w5
        CPidad=1-(w5/w6)**2
        Thetaad=cmath.asin(((w6-w5)/2)/Lad)
        EFFad=1/(1+(1.2*cfad.get()*(((ARad**2)-1)/(2*cmath.tan(Thetaad)))))
        CPactad=EFFad*CPidad
        # static pressure at outlet of axial diffuser
        P6=P5+(CPactad*(Po5-P5))
        #Static enthalpy rise in AD
        Delhad=((cp.get()*T5)/EFFad)*(((P6/P5)**(G3))-1)
        T6=(Delhad/cp.get())+T5
        Po6=P6*((To2/T6)**(G/G1))
        a6=cmath.sqrt(G*R*T6)
        rho6=P6/(R*T6)
        m6=cmath.sqrt((1/G2)*((To2/T6)-1))
        C6=m6*a6
        # losses in axial diffuser
        DHad=To2*cp.get()*(((P6/Po6)**G3)-((P6/Po5)**G3))
        # Overall efficiency calculation------------------
        TEFFSTT=(cp.get()*to1.get()*(((Po6/po1.get())**(G1/G))-1))/DHact
        err1=abs((EFFSTT-TEFFSTT)/TEFFSTT)
        if err1 <= 0.00001:
            count1=10
            EFFSTT=TEFFSTT
        EFFSTT=TEFFSTT
        x+=1
        print("\n%d"%x)
        print("t02/t01=",To2/to1.get())
        print("po6/po1=",Po6/po1.get())
        print("effstt=",EFFSTT)
    print(EFFSTT)
    EFFSTS=(((P6/po1.get())**(G1/G))-1)/((To2/to1.get())-1)
    print(EFFSTS)
    P06P01=Po6/po1.get()
    print(P06P01)
    LID2=LI/d2_
    print(LID2)
    #OEFF1=(DHimp-(DHsf+DHbl+DHtc+DHvld+DHrvd+DHad))/(DHact)
    #DEFFvld=DHvld/DHact
    #DEFFrvd=DHrvd/DHact
    #DEFFad=DHad/DHact
    #DROPEFF=DEFFbl+DEFFtc+DEFFsf+DEFFvld+DEFFrvd+DEFFad
    deffbl.set((DHbl/DHact).real)
    defftc.set((DHtc/DHact).real)
    deffsf.set((DHsf/DHact).real)
    deffdf.set((DHdf/DHact).real)
    deffrcl.set((DHrcl/DHact).real)
    defflk.set((DHlk/DHact).real)
    beta2.set(cmath.acos(C2m/W2).real*180/pi)
    M2.set(M2_.real)
    M2U.set((U2/ao1).real)
    b2.set(b2_.real*1000)
    d2.set(d2_.real*1000)
    d1sd2.set(d1s/d2.get())
    u2.set(U2.real)
    c2u.set(C2u.real)
    c2.set(C2.real)
    c2m.set(C2m.real)
    w2.set(W2.real)
    alpha2.set(Alpha2.real*180/pi)
    effi.set(EFFI.real*100)
    dhimp.set(DHimp.real)
    dhrcl.set(DHrcl.real)
    dhdf.set(DHdf.real)
    c2mu2.set(C2mU2.real)
    effstt.set(EFFSTT.real*100)
    effsts.set(EFFSTS.real*100)
    lid2.set(LID2.real)
    W3.set(w3.real*1000)
    LWRATIO.set((Lrvd/w3).real)
    M4.set(m4.real)
    EFFRVD.set(EFFrvd.real*100)
    ARRVD.set(ARrvd.real)
    W5.set(w5.real*1000)
    M6.set(m6.real)
    ARAD.set(ARad.real)
    EFFAD.set(EFFad.real*100)
    cpactrvd.set(CPactrvd.real)
    cpactad.set(CPactad.real)


def impeller():
    #from cmath import *
    # Input Fluid properties
    Nu=1.81*10**(-5)
    Cv= cp.get()-r.get()
    G=1.4
    G1=(G-1)
    G2=G1/2
    G3=G1/G
    pi=22/7
    # Input Compressor details from cycle
    EFFS=0.80
    # Inputs for inducer calculations
    alpha1=0
    M1rs=m1rs.get()
    N=n.get()
    tb1h=1.8
    tb1s=1
    ZI=16
    Kb1=kb1.get()
    omega=(2*pi*N)/60
    # Inputs for impeller outlet design
    B2b=b2b.get()*pi/180
    W2W1s=w2w1s.get()
    Tipc=tipc.get()
    Cf=cf.get()
    Kb2=kb2.get()
    # beta1s calculation
    #b1s=(((3+(G*M1rs**2))/(2*M1rs**2))*(1-((1-((4*M1rs**2)/(3+(G*M1rs**2))**2))**0.5)))**0.5
    B1s=b1s.get()*pi/180
    # f(M1rt,B1s) calculation
    CB1s=cmath.cos(B1s)
    SB1s=cmath.sin(B1s)
    # r1h/r1s calculation
    ao1=cmath.sqrt(G*r.get()*to1.get())
    #Throat Area calculation to be completed
    s1h=(pi*d1h.get())/ZI
    s1s=(pi*d1s.get())/ZI
    # Inlet state conditions calculations
    rhoo1=po1.get()/(r.get()*to1.get())
    M1term=1+(G2*(M1.get()**2))
    a1=ao1/(M1term**0.5)
    rho1=rhoo1/(M1term**2.5)
    T1=to1.get()/M1term
    P1=po1.get()*((T1/to1.get())**(G/G1))
    # Inlet velocity triangles calculation
    d1rms=(((d1s.get()**2)+(d1h.get()**2))/2)**0.5
    r1rms=d1rms/(2*1000)
    r1mean=(r1s.get()+r1h.get())/2
    C1m=M1.get()*a1
    U1s=(2*pi*r1s.get()*N)/60
    U1rms=(2*pi*r1rms*N)/60
    U1h=(2*pi*r1h.get()*N)/60
    W1s=cmath.sqrt((U1s**2)+(C1m**2))
    W1rms=cmath.sqrt((U1rms**2)+(C1m**2))
    W1h=cmath.sqrt((U1h**2)+(C1m**2))
    B1rms=cmath.acos(C1m/W1rms)
    B1h=cmath.acos(C1m/W1h)
    M1rrms=M1.get()/cmath.cos(B1rms)
    M1rh=M1.get()/cmath.cos(B1h)
    b1=r1s.get()-r1h.get()
    # compressor enthalpy calculation
    DHid=cp.get()*to1.get()*((pr.get()**G3)-1)
    DHact=DHid/EFFS
    To2=to1.get()+(DHact/cp.get())
    # Impeller exit state calculation begins from here
    rho2est=rho1*(To2/to1.get())**(1/G1)
    rho2=rho2est
    DHimp=DHact
    #print("DHimp=",DHimp)
    #slipfactor calculation
    CslipU2=cmath.sqrt(cmath.cos(B2b))/(ZI**0.7)
    # Delta H impeller calculations loop
    count2=0
    i=0
    while count2 < 1:
    #W2/W1s condition check
        C2mU2=0
        W2W1scal=0.00
        while W2W1scal <= W2W1s :
            C2mU2=C2mU2+0.00001
            U2=cmath.sqrt(DHimp/(1-CslipU2-(C2mU2*cmath.tan(B2b))))
            #print("i,C2mU2",i,C2mU2)
            C2m=C2mU2*U2
            C2u=DHimp/U2
            W2u=U2-C2u
            W2=cmath.sqrt((W2u**2)+(C2m**2))
            W2W1scal=W2/W1s
            C2=cmath.sqrt((C2u**2)+(C2m**2))
            alpha2_=cmath.acos(C2m/C2)
            d2_=(U2*60)/(pi*N)
            r2=d2_/2
            Re2=(U2*(d2_/2))/Nu
            # Disk friction loss calculation
            DHdf=(0.01356*rho2*(U2**3)*(d2_**2))/(m.get()*(Re2**0.2))
            # Diffusion factor calculation
            DHimpU22=DHimp/(U2**2)
            D=1-W2W1s+(0.75*DHimpU22*W2W1s*(((ZI/pi)*(1-(r1s.get()/r2))+(2*r1s.get()/r2))**-1))
            # Recirculation loss calculation
            DHrcl=0.02*cmath.sqrt(cmath.tan(alpha2_))*(D**2)*(U2**2)
            # Leakage loss calculation
            b2_=m.get()/(pi*d2_*rho2*C2m*Kb2) # impeller exit width (b2) calculation
            Rc=((2*r2)-(r1h.get()+r1s.get()))/2 # radius of curvature calculation
            Lm=(90/360)*2*pi*Rc  #mean strem line length
            Rbar=(r1mean+r2)/2
            bbar=(b1+b2_)/2
            DPcl=(m.get()*(r2*C2u))/(ZI*Rbar*bbar*Lm)
            #print("Dpcl=",DPcl)
            #print("rho2 second time=",rho2)
            #print("x2=",(2*DPcl)/rho2)
            Ucl=0.816*cmath.sqrt((2*DPcl)/rho2)

            mcl=rho2*ZI*(Tipc/1000)*Lm*Ucl
            DHlk=(mcl*Ucl*U2)/(2*m.get())
            # New Delta H impeller calculation
            DHimp=DHact-DHdf-DHrcl-DHlk
            complex1=cmath.polar(W2W1scal)
            W2W1scal=complex1[0]
        # Diffusion and blade loading loss (Jansen and Coppage)
        DHbl=0.05*(D**2)*(U2**2)
        # Tip clearence loss calculations (Jansen)
        DHtc=(U2**2)*(0.6*((Tipc/1000)/b2_)*C2mU2*cmath.sqrt(((4*pi)/(b2_*ZI))*(((r1s.get()**2)-(r1h.get()**2))/((r2-r1s.get())*(1+(rho2/rho1))))*(C2u/U2)*(C1m/U2)))
        # Impeller Skin friction loss calculations starts
        # Impeller axial length calculation
        LI=d2_*cmath.sqrt(0.28*(M1rs+0.8)*(1-(r1rms/r2))*((r1s.get()-r1h.get())/r2))
        # Impeller Skin friction loss calculations cont..
        Lh=(pi/4)*((LI-(b2_/2))+(r2-r1s.get()-(b1/2)))
        Dh=0.5*(((4*pi*r2*b2_)/((2*pi*r2)+(ZI*b2_)))+((2*pi*((r1s.get()**2)-(r1h.get()**2)))/((pi*(r1s.get()-r1h.get()))+(ZI*b1))))
        Wavg=(W2+((W1s+W1h)/2))/2
        DHsf=((4*Cf*Lh*(Wavg**2))/(2*Dh))
        # Impeller efficiency calcultaions
        EFFI=(DHimp-(DHbl+DHtc+DHsf))/DHimp
        # Density calculation
        Po2=po1.get()*((EFFI*((To2/to1.get())-1)+1)**(G/G1))
        T2=To2-((C2**2)/(2*cp.get()))
        P2=Po2*((T2/To2)**(G/G1))
        # New density calculations
        temp2=P2/(r.get()*T2)
        err2=rho2-temp2
        complex2=cmath.polar(err2)
        err2=complex2[0]
        if err2 < 0.00001:
            count2=10
            rho2=temp2
        rho2=temp2
        i+=1
    a2=cmath.sqrt(G*r.get()*T2)
    deffbl.set((DHbl/DHact).real)
    defftc.set((DHtc/DHact).real)
    deffsf.set((DHsf/DHact).real)
    deffdf.set((DHdf/DHact).real)
    deffrcl.set((DHrcl/DHact).real)
    defflk.set((DHlk/DHact).real)
    beta2.set(cmath.acos(C2m/W2).real*180/pi)
    M2.set((C2/a2).real)
    M2U.set((U2/ao1).real)
    b2.set(b2_.real*1000)
    d2.set(d2_.real*1000)
    d1sd2.set(d1s.get()/d2.get())
    u2.set(U2.real)
    c2u.set(C2u.real)
    c2.set(C2.real)
    c2m.set(C2m.real)
    w2.set(W2.real)
    alpha2.set(alpha2_.real*180/pi)
    effi.set(EFFI.real)
    dhimp.set(DHimp.real)
    dhrcl.set(DHrcl.real)
    dhdf.set(DHdf.real)
    c2mu2.set(C2mU2.real)
    lid2.set((LI/d2_).real)


def setrv(ao1,gamma,rv):
    a1=ao1
    rho=po1.get()/r.get()/to1.get()
    rho1=rho
    r1h_=0.01
    temp=0
    temprho=rho1
    min=100
    i=0
    while i<2000:
        r1h_=0.01
        min=100
        while temp<1:
            cons1=(kb1.get()*pi*rho1*r1h_*r1h_*(1/rv/rv-1))
            tempmw1t=1/a1*sqrt(pow(m.get()/cons1,2)+pow(2*pi*n.get()*r1h_/60/rv-m.get()*tan(alpha1.get())/cons1,2))
            if tempmw1t <= min:
                r1h_+=0.00001
                mw1t=tempmw1t
                min=mw1t
            else:
                r1h_-=0.00001
                r1h_=round(r1h_,5)
                break
        r1t=r1h_/rv
        u1t=2*pi*n.get()*r1t/60
        v1a=m.get()/(kb1.get()*pi*(r1t*r1t-r1h_*r1h_)*rho1)
        w1t=sqrt(v1a*v1a+pow(u1t-v1a*tan(alpha1.get()),2))
        b1s_=acos(v1a/w1t)
        m1=mw1t*cos(b1s_)
        a1=ao1/sqrt(1+(gamma-1)/2*m1*m1)
        rho1=rho/pow(1+(gamma-1)/2*m1*m1,2.5)
        if abs(temprho-rho1)>0.0001:
            temprho=rho1
            i+=1
        else :
            print("iterations done!!")
            break
        if i==1999:
            print("iterations exhausted!")

    c1m.set(v1a)
    b1s.set(b1s_*180/pi)
    m1rs.set(mw1t)
    r1h.set(r1h_)
    r1s.set(r1h.get()/rv)
    M1.set(m1)

    print("\n")
    print("rho1=\t",rho1)
    print("rho=\t",rho)
    print("temprho-rho1=\t",temprho-rho1)
    print("a1=\t",a1)
    print("ao1=\t",ao1)


def setr1h(ao1,gamma,r1h):
    a1=ao1
    rho=po1.get()/r.get()/to1.get()
    rho1=rho
    rv_=0.75#maximum
    temp=0
    temprho=rho1
    min=100
    i=0
    while i<2000:
        rv_=0.75
        min=100
        while temp<1:
            cons1=(kb1.get()*pi*rho1*r1h*r1h*(1/rv_/rv_-1))
            tempmw1t=1/a1*sqrt(pow(m.get()/cons1,2)+pow(2*pi*n.get()*r1h/60/rv_-m.get()*tan(alpha1.get())/cons1,2))
            if tempmw1t <= min:
                rv_-=0.001
                mw1t=tempmw1t
                min=mw1t
            else:
                rv_+=0.001
                rv_=round(rv_,3)
                break
        r1t=r1h/rv_ #r1s
        u1t=2*pi*n.get()*r1t/60 #u1s
        v1a=m.get()/(kb1.get()*pi*(r1t*r1t-r1h*r1h)*rho1)
        w1t=sqrt(v1a*v1a+pow(u1t-v1a*tan(alpha1.get()),2))
        b1s_=acos(v1a/w1t)
        m1=mw1t*cos(b1s_)
        a1=ao1/sqrt(1+(gamma-1)/2*m1*m1)
        rho1=rho/pow(1+(gamma-1)/2*m1*m1,2.5)
        if abs(temprho-rho1)>0.0001:
            temprho=rho1
            i+=1
        else :
            print("iterations done!!")
            break
        if i==1999:
            print("iterations exhausted!")

    c1m.set(v1a)
    b1s.set(b1s_*180/pi)
    m1rs.set(mw1t)
    rv.set(rv_)
    r1s.set(r1h/rv_)
    M1.set(m1)

    print("\n")
    print("rho1=\t",rho1)
    print("rho=\t",rho)
    print("error %=\t",(temprho-rho1)/rho1*100)
    print("a1=\t",a1)
    print("ao1=\t",ao1)


def plotfigure():
    #canvas=Canvas(page6,width=800,height=480,bg='white')
    #canvas.grid(row=0,column=0)

    f = Figure(figsize=(8,6), dpi=100)
    a = f.add_subplot(111,aspect=1)

    h=c1m.get()
    a.axis('off')
    a.plot([10,10],[10,h+10])
    a.text(-55,10+h/2,"C1m=\n%s"%round(c1m.get(),2))
    a.add_patch(Arc((10,10),80,80,angle=0,theta1=90-b1s.get(),theta2=90))
    a.text(30,10+h/3,"%s"%(round(b1s.get(),2))+u"\u00B0")
    a.plot([10,u1s.get()+10],[h+10,h+10])
    a.plot([10,u1s.get()+10],[10,h+10])
    a.text(u1s.get()/2-10,-30+h/2,'W1s= %s'%round(w1s.get(),2))

    a.plot([40+u1s.get(),40+u1s.get()],[10,h+10])
    a.plot([40+u1s.get(),u1rms.get()+40+u1s.get()],[h+10,h+10])
    a.plot([40+u1s.get(),u1rms.get()+40+u1s.get()],[10,h+10])
    a.text(u1rms.get()/2+30+u1s.get()-15,-30+h/2,'W1rms= %s'%round(w1rms.get(),2))
    a.add_patch(Arc((40+u1s.get(),10),80,80,angle=0,theta1=90-b1rms.get(),theta2=90))
    a.text(50+u1s.get(),10+h/3,"%s"%(round(b1rms.get(),2))+u"\u00B0")

    a.plot([u1s.get()+u1rms.get()+80,u1s.get()+u1rms.get()+80],[10,h+10])
    a.plot([u1s.get()+u1rms.get()+80,u1h.get()+u1s.get()+u1rms.get()+80],[h+10,h+10])
    a.plot([u1s.get()+u1rms.get()+80,u1h.get()+u1s.get()+u1rms.get()+80],[10,h+10])
    a.text(u1h.get()/2+u1s.get()+u1rms.get()+100,10+h/2,'W1h= %s'%round(w1h.get(),2))
    a.add_patch(Arc((u1s.get()+u1rms.get()+80,10),80,80,angle=0,theta1=90-b1h.get()
    ,theta2=90))
    a.text(u1s.get()+u1rms.get()+85,15+h/3,"%s"%(round(b1h.get(),2))+u"\u00B0",size='small')

    global canvas
    canvas = FigureCanvasTkAgg(f,page6)
    canvas.draw()
    canvas.get_tk_widget().grid(row=0,column=0)

    base=h+100
    h2=c2m.get()

    a.plot([10,u2.get()+10],[h2+base,h2+base])
    a.text(u2.get()/2+10,h2+base+10,'U2= %s'%round(u2.get(),2))
    a.plot([10,c2u.get()+10],[h2+base,h2+base])
    a.plot([10,c2u.get()+10],[h2+base,base])
    a.text(c2u.get()/2-30,base+h2/2-30,'C2=\n%s'%round(c2.get(),2))
    a.plot([c2u.get()+10,u2.get()+10],[base,h2+base])
    a.text(c2u.get()+10+(u2.get()-c2u.get())/2+10,base+h2/2-10,'W2= %s'%round(w2.get(),2))
    a.plot([c2u.get()+10,c2u.get()+10],[base,base+h2])
    a.text(c2u.get()-50,base+h2/2+10,'C2m=\n%s'%round(c2m.get(),2))
    a.add_patch(Arc((c2u.get()+10,base),50,50,angle=0,theta1=90,theta2=90+alpha2.get()))
    a.text(c2u.get()-50,base+h/4,"%s"%(round(alpha2.get(),2))+u"\u00B0")
    a.add_patch(Arc((c2u.get()+10,base),70,70,angle=0,theta2=90,theta1=90-beta2.get()))
    a.text(c2u.get()+20,base+h/4,"%s"%(round(beta2.get(),2))+u"\u00B0",size='small')

    a.text(550,base+h2+100,'*All units in m/s')

def calculate2(m1,t1):
    d1h=r1h.get()*2
    d1s=r1s.get()*2
    N=n.get()
    tempd1r=sqrt((d1h*d1h+d1s*d1s)/2)
    d1rms.set(tempd1r)
    u1s.set(pi*d1s*N/60)
    u1rms.set(pi*d1rms.get()*N/60)
    u1h.set(pi*d1h*N/60)
    w1s.set(sqrt(u1s.get()*u1s.get()+c1m.get()*c1m.get()))
    w1rms.set(sqrt(u1rms.get()*u1rms.get()+c1m.get()*c1m.get()))
    w1h.set(sqrt(u1h.get()*u1h.get()+c1m.get()*c1m.get()))
    temp=acos(c1m.get()/w1rms.get())
    b1rms.set(temp)
    temp1=acos(c1m.get()/w1h.get())
    b1h.set(temp1)
    to1rel.set(t1+w1rms.get()*w1rms.get()/(2*cp.get()))
    m1r_rms.set(m1/cos(b1rms.get()))
    m1rh.set(m1/cos(b1h.get()))
    entryb1h.delete(0,END)
    entryb1h.insert(END,temp1*180/pi)
    entryb1r.delete(0,END)
    entryb1r.insert(END,temp*180/pi)
    entryd1h.delete(0,END)
    entryd1h.insert(END,d1h*1000)
    entryd1r.delete(0,END)
    entryd1r.insert(END,tempd1r*1000)
    entryd1s.delete(0,END)
    entryd1s.insert(END,d1s*1000)


def calcb1s(m1r,gamma):
    temp1=1-4*m1r*m1r/(pow(3+gamma*m1r*m1r,2))
    temp2=1-sqrt(temp1)
    temp3=(3+gamma*m1r*m1r)/(2*m1r*m1r)*temp2
    return(acos(sqrt(temp3))*180/pi)

def calcfunc(m1r,gamma,b):
    temp1=(1+(gamma-1)/2*m1r*m1r*pow(cos(b),2))
    temp2=m1r*m1r*m1r*pow(sin(b),2)*cos(b)/(pow(temp1,(3*gamma-1)/2/(gamma-1)))
    func.set(temp2)

def calck(mdot,kb,N,gamma,f,p1,ao1):
    omega=2*pi*N/60
    k.set(mdot*omega*omega/kb/gamma/pi/f/p1/ao1)

def calr1s(K,r1H):
    r1s.set(r1H/sqrt(1-K))
    d1s.set(r1s.get()*2)

rows=0
while rows<50:
    win.rowconfigure(rows,weight=1)
    win.columnconfigure(rows,weight=1)
    rows+=1

nb=ttk.Notebook(win)
nb.grid(row=1,column=0, rowspan=40, columnspan=50, sticky='NESW')

buttoncalc=Button(win,text="calculate",command=calculate)
buttoncalc.grid(row=45,column=0)

Button(win, text="Export Data",command=export,padx=2).grid(row=45,column=2)


#*****page1*******
page1=ttk.Frame(nb)
nb.add(page1,text="Inputs")

Label(page1,text="Air Properties",font="helvetica 10 bold",pady=2).grid(row=2,column=0,columnspan=2)

labelcv=Label(page1,text="CV(J/kg/k)")
labelcv.grid(row=3,column=0)

labelcp=Label(page1,text="CP(J/kg/K)")
labelcp.grid(row=4,column=0)

labelr=Label(page1,text="R(J/kg/K)")
labelr.grid(row=5,column=0)

cv=DoubleVar()
cv.set(718)
entrycv=Entry(page1,textvariable=cv)
entrycv.grid(row=3,column=1)

cp=DoubleVar()
cp.set(1005)
entrycp=Entry(page1,textvariable=cp)
entrycp.grid(row=4,column=1)

r=DoubleVar()
r.set(287)
entryr=Entry(page1,textvariable=r)
entryr.grid(row=5,column=1)

Label(page1,text="").grid(row=6,column=3)
ttk.Separator(page1,orient='horizontal').grid(row=7,columnspan=7,sticky='ew')

#****Cycle Parameters*****
Label(page1,text="Cycle Parameters",font="helvetica 10 bold").grid(row=8,column=0,columnspan=2)
labelpr=Label(page1,text="Pressure ratio")
labelpr.grid(row=9,column=0)

labelm=Label(page1,text="Mass flow(kg/s)")
labelm.grid(row=10,column=0)

labelTo1=Label(page1,text="To1 (K)")
labelTo1.grid(row=11,column=0)

labelPo1=Label(page1,text="Po1 (Pascal)")
labelPo1.grid(row=12,column=0)

labelN=Label(page1,text="N(rpm)")
labelN.grid(row=13,column=0)

pr=DoubleVar()
pr.set(4)
entrypr=Entry(page1,textvariable=pr)
entrypr.grid(row=9,column=1)

m=DoubleVar()
m.set(1.187)
entrym=Entry(page1,textvariable=m)
entrym.grid(row=10,column=1)

to1=DoubleVar()
to1.set(300)
entryto1=Entry(page1,textvariable=to1)
entryto1.grid(row=11,column=1)

po1=DoubleVar()
po1.set(101325)
entrypo1=Entry(page1,textvariable=po1)
entrypo1.grid(row=12,column=1)

n=DoubleVar()
n.set(44000)
entryn=Entry(page1,textvariable=n)
entryn.grid(row=13,column=1)

Label(page1,text="").grid(row=14)
ttk.Separator(page1,orient='horizontal').grid(row=15,columnspan=7,sticky='ew')

#****Inlet Parameters*****

Label(page1,text="Inlet Parameters",font="helvetica 10 bold").grid(row=16,columnspan=2)
labelalpha1=Label(page1,text="alpha 1")
labelalpha1.grid(row=21,column=0)

labelm1rs=Label(page1,text="Relative mach number\n shroud")
labelm1rs.grid(row=17,column=0)

labelr1h=Label(page1,text="Hub radius (m)")
labelr1h.grid(row=18,column=0)

labelrv=Label(page1,text="hub/shroud radius\n ratio")
labelrv.grid(row=19,column=0)

labelkb1=Label(page1,text="Blockage Factor")
labelkb1.grid(row=20,column=0)

alpha1=DoubleVar()
entryalpha1=Entry(page1,textvariable=alpha1)
entryalpha1.grid(row=21,column=1)


def changebg(*args):
    if radiom1rs.get()==0:
        entrym1rs.config(state="normal")
        menu.config("disabled")
        entryrv.config(state="disabled")
        entryr1h.config(state="normal")

    else:
        entrym1rs.config(state="disabled")
        menu.config(state="normal")
        if radiom1rs.get()==1 and menum1rs.get()=="Set rv":
            entryrv.config(state="normal")
            entryr1h.config(state="disabled")
        else:
            entryrv.config(state="disabled")
            entryr1h.config(state="normal")
Label(page1,text="").grid(row=22)

radiom1rs=IntVar()
Radiobutton(page1,text="specify mach number",value=0,variable=radiom1rs,command=changebg).grid(row=23,column=0)
Radiobutton(page1,text="optimize",value=1,variable=radiom1rs,command=changebg).grid(row=23,column=1)

menum1rs=StringVar()
menum1rs.set('Set r1h')
menu=OptionMenu(page1,menum1rs,'Set r1h','Set rv')
menu.grid(row=23,column=3)
menu.config(state="disabled",padx=2)
menum1rs.trace('w',changebg)
radiom1rs.set(0)

m1rs=DoubleVar()
m1rs.set(0.85)
entrym1rs=Entry(page1,textvariable=m1rs)
entrym1rs.grid(row=17,column=1)

r1h=DoubleVar()
r1h.set(0.0225)
entryr1h=Entry(page1,textvariable=r1h)
entryr1h.grid(row=18,column=1)

rv=DoubleVar()
rv.set(0.42)
entryrv=Entry(page1,textvariable=rv)
entryrv.config(state="disabled")
entryrv.grid(row=19,column=1)


kb1=DoubleVar()
kb1.set(0.975)
entrykb1=Entry(page1,textvariable=kb1)
entrykb1.grid(row=20,column=1)

ttk.Separator(page1,orient='vertical').grid(column=12,row=3,rowspan=20,sticky='ns',padx=8)

#*****Exit Parameters*****

Label(page1,text="Exit Parameters",font="helvetica 10 bold").grid(row=2,column=13,columnspan=2)

def togglealpha2(*args):
    entries=[("rr32"),("cfvld"),("r4r3"),("zrd"),
    ("cfrvd"),("lwratio"),("zad"),("cfad"),("effstt"),("effsts"),
    ("W3"),"LWRATIO","M4","ARRVD","EFFRVD","W5","M6","ARAD","EFFAD","cpactrvd","cpactad"]
    for name in entries:
        if menuw2w1s.get()=='Impeller with diffuser':
            exec("entry""%s.config(state='normal')"%name)
        else:
            exec("entry""%s.config(state='disabled')"%name)

Label(page1,text="B2b (deg)").grid(row=3,column=13)
Label(page1,text="W2/W1s").grid(row=4,column=13)
Label(page1,text="Tip clearance").grid(row=5,column=13)
Label(page1,text="Cf").grid(row=6,column=13)
Label(page1,text="Blockage Factor 2").grid(row=7,column=13)

b2b=DoubleVar()
entryb2b=Entry(page1,textvariable=b2b)
entryb2b.grid(row=3,column=14)
b2b.set(30)

w2w1s=DoubleVar()
entryw2w1s=Entry(page1,textvariable=w2w1s)
entryw2w1s.grid(row=4,column=14)
w2w1s.set(0.62)

tipc=DoubleVar()
entrytipc=Entry(page1,textvariable=tipc)
entrytipc.grid(row=5,column=14)
tipc.set(0.4)

cf=DoubleVar()
entrycf=Entry(page1,textvariable=cf)
entrycf.grid(row=6,column=14)
cf.set(0.04)

kb2=DoubleVar()
entrykb2=Entry(page1,textvariable=kb2)
entrykb2.grid(row=7,column=14)
kb2.set(0.965)

alpha2=DoubleVar()

row_=10
variable_list=[("Vaneless space \n radius ratio","rr32"),("cfvld","cfvld"),("r4r3","r4r3"),("Zrd","zrd"),
("cfrvd","cfrvd"),("L/W ratio","lwratio"),("Zad","zad"),("cfad","cfad")]
for name,var in variable_list:
    exec("Label(page1,text=name).grid(row=row_,column=13)")
    exec("%s=DoubleVar()"%var)
    exec("entry""%s=Entry(page1,textvariable=%s,state='disabled')"%(var,var))
    exec("entry""%s.grid(row=row_,column=14,pady=2)"%var)
    row_+=1

rr32.set(1.1)
cfvld.set(0.03)
r4r3.set(1.25)
zrd.set(16)
cfrvd.set(0.03)
lwratio.set(4)
zad.set(48)
cfad.set(0.03)

menuw2w1s=StringVar()
menuw2w1s.set('Only Impeller')
menu2=OptionMenu(page1,menuw2w1s,'Only Impeller','Impeller with diffuser')
menu2.grid(row=9,column=15)
menuw2w1s.trace('w',togglealpha2)


#*****results****
page3=ttk.Frame(nb)
nb.add(page3,text="Results")


func=DoubleVar()
k=DoubleVar()
r1s=DoubleVar()

#giving a empty row as seperation
Label(page3,text="Impeller entry",font="helvetica 9 bold").grid(row=5,column=3)

Label(page3,text="Values(mm)").grid(row=6,column=1)
Label(page3,text="U1(m/s)").grid(row=6,column=2)
Label(page3,text="W1(m/s)").grid(row=6,column=3)
Label(page3,text=u"\u03B2""1(deg)").grid(row=6,column=4)
Label(page3,text="M1 rel").grid(row=6,column=5)
Label(page3,text="C1m(m/s)").grid(row=6,column=6)
Label(page3,text="M1").grid(row=12,column=0)
Label(page3,text="Radius ratio").grid(row=12,column=3)

d1s=DoubleVar()
entryd1s=Entry(page3,textvariable=d1s)
entryd1s.grid(row=7,column=1)

d1rms=DoubleVar()
entryd1r=Entry(page3,textvariable=d1rms)
entryd1r.grid(row=8,column=1)

d1h=DoubleVar()
d1h.set(r1h.get()*2)
entryd1h=Entry(page3,textvariable=d1h)
entryd1h.grid(row=9,column=1)

u1s=DoubleVar()
Entry(page3,textvariable=u1s).grid(row=7,column=2)

u1rms=DoubleVar()
Entry(page3,textvariable=u1rms).grid(row=8,column=2)

u1h=DoubleVar()
Entry(page3,textvariable=u1h).grid(row=9,column=2)

M1=DoubleVar()
Entry(page3,textvariable=M1).grid(row=12,column=1)

Entry(page3,textvariable=rv).grid(row=12,column=4)

w1s=DoubleVar()
Entry(page3,textvariable=w1s).grid(row=7,column=3)

Label(page3,text="d1s").grid(row=7,column=0)
Label(page3,text="d1rms").grid(row=8,column=0)
Label(page3,text="d1h").grid(row=9,column=0)
Label(page3,text="").grid(row=11,column=0)

w1rms=DoubleVar()
Entry(page3,textvariable=w1rms).grid(row=8,column=3)

w1h=DoubleVar()
Entry(page3,textvariable=w1h).grid(row=9,column=3)

b1s=DoubleVar()
entryb1s=Entry(page3,textvariable=b1s)
entryb1s.grid(row=7,column=4)

b1rms=DoubleVar()
entryb1r=Entry(page3,textvariable=b1rms)
entryb1r.grid(row=8,column=4)

b1h=DoubleVar()
entryb1h=Entry(page3,textvariable=b1h)
entryb1h.grid(row=9,column=4)

Entry(page3,textvariable=m1rs).grid(row=7,column=5)

m1r_rms=DoubleVar()
Entry(page3,textvariable=m1r_rms).grid(row=8,column=5)

m1rh=DoubleVar()
Entry(page3,textvariable=m1rh).grid(row=9,column=5)

c1m=DoubleVar()
Entry(page3,textvariable=c1m).grid(row=7,column=6)
Entry(page3,textvariable=c1m).grid(row=8,column=6)
Entry(page3,textvariable=c1m).grid(row=9,column=6)


to1rel=DoubleVar()
    #*****impeller exit*****
Label(page3,text="").grid(row=13,column=0)
ttk.Separator(page3,orient='horizontal').grid(row=14,columnspan=10,sticky='ew')
Label(page3,text="Impeller Exit",font='helvetica 9 bold').grid(row=15,column=3)


list1=["DEFFbl","DEFFtc","DEFFsf","DEFFdf","DEFFrcl","DEFFlk"]
col=0
for name in list1:
    exec("Label(page3,text=name).grid(row=23,column=col)")
    exec("%s=DoubleVar()"%name.lower())
    exec("Entry(page3,textvariable=%s).grid(row=24,column=col)"%name.lower())
    col+=1

list2=[("b2(mm)","b2"),("d2(mm)","d2"),("U2(m/s)","u2"),("C2U(m/s)","c2u"),("C2(m/s)","c2"),
("C2m(m/s)","c2m"),("W2(m/s)","w2")]
col=0
for label,var in list2:
    exec("Label(page3,text=label).grid(row=16,column=col)")
    exec("%s=DoubleVar()"%var)
    exec("Entry(page3,textvariable=%s).grid(row=17,column=col)"%var)
    col+=1

list3=[("EFFI","effi"),("M2","M2"),("M2U","M2U"),(u"\u03B1""2(deg)","alpha2"),
(u"\u03B2""2(deg)","beta2"),("d1s/d2","d1sd2"),("LI/d2","lid2")]
col=0
for label,var in list3:
    exec("Label(page3,text=label).grid(row=18,column=col)")
    exec("%s=DoubleVar()"%var)
    exec("Entry(page3,textvariable=%s).grid(row=19,column=col)"%var)
    col+=1
Label(page3,text="").grid(row=20)
ttk.Separator(page3,orient='horizontal').grid(row=21,columnspan=10,sticky='ew')
Label(page3,text="Impeller overall",font='helvetica 9 bold').grid(row=22,column=3)
c2mu2=DoubleVar()

Label(page3,text="DHimp(J)").grid(row=25,column=0)
dhimp=DoubleVar()
Entry(page3,textvariable=dhimp).grid(row=25,column=1)

Label(page3,text="DHrcl(J)").grid(row=25,column=2)
dhrcl=DoubleVar()
Entry(page3,textvariable=dhrcl).grid(row=25,column=3)

Label(page3,text="DHdf(J)").grid(row=25,column=4)
dhdf=DoubleVar()
Entry(page3,textvariable=dhdf).grid(row=25,column=5)

Label(page3,text="").grid(row=26)
ttk.Separator(page3,orient='horizontal').grid(row=27,columnspan=10,sticky='ew',pady=2)

Label(page3,text="Radial Difuuser",font="helvetical 10 bold").grid(row=28,column=0,columnspan=2)
listrad=[("width(mm)","W3"),("Len/width","LWRATIO"),("Area Ratio","ARRVD"),("M4","M4"),("effi","EFFRVD"),
("CP actual","cpactrvd")]
column_=0
for name,var in listrad:
    exec("Label(page3,text=name).grid(row=29,column=column_)")
    exec("%s=DoubleVar()"%var)
    exec("entry""%s=Entry(page3,textvariable=%s,state='disabled')"%(var,var))
    exec("entry""%s.grid(column=column_,row=30)"%var)
    column_+=1

Label(page3,text="Axial Difuuser",font="helvetical 10 bold").grid(row=31,column=0,columnspan=2)
listaxi=[("width(mm)","W5"),("Area Ratio","ARAD"),("M6","M6"),("effi","EFFAD"),("CP actual","cpactad")]
column_=0
for name,var in listaxi:
    exec("Label(page3,text=name).grid(row=32,column=column_)")
    exec("%s=DoubleVar()"%var)
    exec("entry""%s=Entry(page3,textvariable=%s,state='disabled')"%(var,var))
    exec("entry""%s.grid(column=column_,row=33)"%var)
    column_+=1

Label(page3,text="Overall Compressor",font="helvetical 10 bold").grid(row=28,column=6)

var_list=[("effstt"),("effsts")]
row_=29
for var in var_list:
    exec("Label(page3,text=var.upper()).grid(row=row_,column=6)")
    exec("%s=DoubleVar()"%var)
    exec("entry""%s=Entry(page3,textvariable=%s,state='disabled' if var!='lid2' else 'normal')"%(var,var))
    exec("entry""%s.grid(column=6,row=row_+1)"%var)
    row_+=2



#*****page 6*****
page6=ttk.Frame(nb)
nb.add(page6,text="plots")

win.mainloop()
